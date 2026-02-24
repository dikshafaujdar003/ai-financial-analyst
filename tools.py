from langchain.tools import tool
import yfinance as yf
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import os
import re
import json
from datetime import datetime


# ── Helpers ────────────────────────────────────────────────────────────────────
def _clean_ticker(ticker: str) -> str:
    """Strip any parameter name prefixes the LLM might add, e.g. TICKER='AAPL' → AAPL"""
    ticker = ticker.strip().upper()
    ticker = re.sub(r'^[A-Z_]+=["\'"]?', '', ticker)
    ticker = ticker.strip("\"' ")
    return ticker


def _read_file(path: str) -> pd.DataFrame:
    """Read CSV or Excel, return DataFrame."""
    if path.endswith((".xlsx", ".xls")):
        return pd.read_excel(path)
    return pd.read_csv(path)


def _find_file(name: str) -> str | None:
    """Find a file by partial name in current directory."""
    name = name.strip()
    if os.path.exists(name):
        return name
    for f in os.listdir("."):
        if name.lower() in f.lower():
            return f
    return None


# ── Stock Tools ────────────────────────────────────────────────────────────────
@tool
def get_stock_data(ticker: str) -> str:
    """Downloads 1 year of stock data for a given ticker symbol and saves it to a CSV file."""
    ticker = _clean_ticker(ticker)
    df = yf.download(ticker, period="1y")
    if df.empty:
        return f"Error: No data found for {ticker}."
    df.to_csv(f"{ticker}_raw.csv")
    return f"Success: Data saved to {ticker}_raw.csv."


@tool
def clean_and_profile(ticker: str) -> str:
    """Reads the raw CSV for a ticker, cleans the data, and returns a statistical profile."""
    ticker = _clean_ticker(ticker)
    path = f"{ticker}_raw.csv"
    if not os.path.exists(path):
        return f"Error: {path} not found. Run get_stock_data first."

    df = pd.read_csv(path, index_col=0, parse_dates=True, header=[0, 1])
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.dropna(how="all", inplace=True)
    df = df.apply(pd.to_numeric, errors="coerce")
    df.to_csv(f"{ticker}_clean.csv")

    profile  = df["Close"].describe().round(2)
    latest   = round(float(df["Close"].iloc[-1]), 2)
    pct_chg  = round(float(df["Close"].pct_change().iloc[-1]) * 100, 4)

    return (
        f"--- Profile for {ticker} ---\n"
        f"Rows: {len(df)}\nLatest Close: ${latest}\n1-Day Change: {pct_chg:.2f}%\n"
        f"Stats (Close):\n{profile.to_string()}\nCleaned data saved to {ticker}_clean.csv."
    )


@tool
def create_chart(ticker: str) -> str:
    """Creates a closing price chart with 50-day MA for a ticker and saves it as PNG."""
    ticker = _clean_ticker(ticker)
    path = f"{ticker}_clean.csv" if os.path.exists(f"{ticker}_clean.csv") else f"{ticker}_raw.csv"
    if not os.path.exists(path):
        return f"Error: No data for {ticker}. Run get_stock_data first."

    df = pd.read_csv(path, index_col=0, parse_dates=True, header=[0, 1])
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df = df.apply(pd.to_numeric, errors="coerce")
    df.dropna(subset=["Close"], inplace=True)

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.plot(df.index, df["Close"], label="Close Price", color="royalblue", linewidth=1.5)
    if len(df) >= 50:
        ax.plot(df.index, df["Close"].rolling(50).mean(),
                label="50-day MA", color="orange", linewidth=1.2, linestyle="--")
    ax.set_title(f"{ticker} — Closing Price (Last 1 Year)", fontsize=14)
    ax.set_xlabel("Date"); ax.set_ylabel("Price (USD)")
    ax.legend(); ax.grid(True, linestyle="--", alpha=0.5)
    plt.tight_layout()
    chart_path = f"{ticker}_chart.png"
    plt.savefig(chart_path, dpi=150)
    plt.close()
    return f"Success: Chart saved to {chart_path}."


# ── News Sentiment ─────────────────────────────────────────────────────────────
@tool
def search_news(ticker: str) -> str:
    """
    Searches for recent news headlines about a stock ticker and returns
    a sentiment summary (bullish / bearish / neutral) with key headlines.
    """
    from duckduckgo_search import DDGS
    ticker = _clean_ticker(ticker)

    try:
        with DDGS() as ddgs:
            results = list(ddgs.news(f"{ticker} stock news", max_results=8))
    except Exception as e:
        return f"Error fetching news for {ticker}: {e}"

    if not results:
        return f"No recent news found for {ticker}."

    headlines  = [r.get("title", "") for r in results]
    bodies     = " ".join([r.get("body", "") for r in results]).lower()

    bullish_words = ["surge", "soar", "rally", "beat", "strong", "growth",
                     "profit", "upgrade", "buy", "outperform", "record", "gain"]
    bearish_words = ["drop", "fall", "miss", "weak", "loss", "downgrade",
                     "sell", "underperform", "decline", "cut", "risk", "concern"]

    bull_score = sum(bodies.count(w) for w in bullish_words)
    bear_score = sum(bodies.count(w) for w in bearish_words)

    if bull_score > bear_score * 1.3:
        sentiment = "🟢 BULLISH"
    elif bear_score > bull_score * 1.3:
        sentiment = "🔴 BEARISH"
    else:
        sentiment = "🟡 NEUTRAL"

    headline_text = "\n".join(f"  - {h}" for h in headlines[:6])
    return (
        f"--- News Sentiment for {ticker} ---\n"
        f"Sentiment: {sentiment} (bullish signals: {bull_score}, bearish signals: {bear_score})\n\n"
        f"Recent Headlines:\n{headline_text}"
    )


# ── Multi-Stock Comparison ─────────────────────────────────────────────────────
@tool
def compare_stocks(tickers: str) -> str:
    """
    Compares two stocks side by side.
    Pass two ticker symbols separated by a comma, e.g. 'NVDA,TSLA'
    """
    parts = [t.strip().upper() for t in tickers.replace(" ", "").split(",")]
    if len(parts) != 2:
        return "Error: Please provide exactly 2 tickers separated by a comma, e.g. 'NVDA,TSLA'"

    t1, t2 = parts
    frames = {}

    for ticker in [t1, t2]:
        df = yf.download(ticker, period="1y")
        if df.empty:
            return f"Error: No data found for {ticker}."
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df = df.apply(pd.to_numeric, errors="coerce")
        df.dropna(subset=["Close"], inplace=True)
        frames[ticker] = df

    df1, df2 = frames[t1], frames[t2]

    def metrics(df, ticker):
        start      = float(df["Close"].iloc[0])
        end        = float(df["Close"].iloc[-1])
        pct        = round((end - start) / start * 100, 2)
        hi         = round(float(df["Close"].max()), 2)
        lo         = round(float(df["Close"].min()), 2)
        daily_ret  = df["Close"].pct_change().dropna()
        vol        = round(float(daily_ret.std() * 100), 2)
        rf_daily   = 0.05 / 252
        excess     = daily_ret - rf_daily
        sharpe     = round(float(excess.mean() / excess.std() * (252 ** 0.5)), 2) if excess.std() != 0 else 0
        roll_max   = df["Close"].cummax()
        drawdown   = ((df["Close"] - roll_max) / roll_max).min()
        max_dd     = round(float(drawdown) * 100, 2)
        return {"ticker": ticker, "start": round(start, 2), "end": round(end, 2),
                "return": pct, "high": hi, "low": lo,
                "volatility": vol, "sharpe": sharpe, "max_drawdown": max_dd}

    m1, m2 = metrics(df1, t1), metrics(df2, t2)

    # Side-by-side chart
    fig, axes = plt.subplots(1, 2, figsize=(16, 5))
    for ax, df, m in [(axes[0], df1, m1), (axes[1], df2, m2)]:
        ax.plot(df.index, df["Close"], color="royalblue", linewidth=1.5, label="Close")
        if len(df) >= 50:
            ax.plot(df.index, df["Close"].rolling(50).mean(),
                    color="orange", linewidth=1.2, linestyle="--", label="50-day MA")
        color = "green" if m["return"] >= 0 else "red"
        ax.set_title(f"{m['ticker']} | Return: {m['return']:+.2f}%", color=color, fontsize=13)
        ax.set_xlabel("Date"); ax.set_ylabel("Price (USD)")
        ax.legend(); ax.grid(True, linestyle="--", alpha=0.5)

    plt.suptitle(f"{t1} vs {t2} — 1 Year Comparison", fontsize=15, fontweight="bold")
    plt.tight_layout()
    chart_path = f"{t1}_vs_{t2}_comparison.png"
    plt.savefig(chart_path, dpi=150)
    plt.close()

    winner = m1 if m1["return"] > m2["return"] else m2
    loser  = m2 if winner == m1 else m1
    diff   = round(abs(m1["return"] - m2["return"]), 2)

    verdict = (
        f"📊 {t1} vs {t2} — 1 Year Comparison\n\n"
        f"  {t1}: Return={m1['return']:+.2f}% | High=${m1['high']} | Low=${m1['low']} "
        f"| Volatility={m1['volatility']}% | Sharpe={m1['sharpe']} | Max Drawdown={m1['max_drawdown']}%\n"
        f"  {t2}: Return={m2['return']:+.2f}% | High=${m2['high']} | Low=${m2['low']} "
        f"| Volatility={m2['volatility']}% | Sharpe={m2['sharpe']} | Max Drawdown={m2['max_drawdown']}%\n\n"
        f"🏆 Verdict: {winner['ticker']} outperformed {loser['ticker']} by {diff}% over the past year.\n"
    )

    if winner["volatility"] > loser["volatility"] * 1.3:
        verdict += (
            f"⚠️ Note: {winner['ticker']} achieved higher returns but with "
            f"significantly more volatility ({winner['volatility']}% vs {loser['volatility']}%)."
        )

    verdict += f"\nChart saved to {chart_path}."
    return verdict


# ── Financial Modeling ─────────────────────────────────────────────────────────
@tool
def financial_model(assumptions: str) -> str:
    """
    Builds a 3-year what-if financial model and exports it to Excel.
    Pass assumptions as comma-separated key=value pairs, e.g.:
    'revenue=1000000, cost_ratio=0.6, price_increase=0.05, growth_rate=0.1, tax_rate=0.21'
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    params = {}
    try:
        params = json.loads(assumptions)
    except Exception:
        for part in assumptions.replace(";", ",").split(","):
            if "=" in part:
                k, v = part.strip().split("=", 1)
                try:
                    params[k.strip()] = float(v.strip())
                except Exception:
                    pass

    rev0       = float(params.get("revenue", 1_000_000))
    cost_ratio = float(params.get("cost_ratio", 0.60))
    price_inc  = float(params.get("price_increase", 0.05))
    growth     = float(params.get("growth_rate", 0.10))
    tax        = float(params.get("tax_rate", 0.21))

    years      = ["Base Year", "Year 1", "Year 2", "Year 3"]
    revs, cogs, gross, op_profit, net = [], [], [], [], []
    rev = rev0
    for i in range(4):
        if i > 0:
            rev = rev * (1 + growth) * (1 + price_inc)
        c  = rev * cost_ratio
        g  = rev - c
        op = g * 0.80
        n  = op * (1 - tax)
        revs.append(rev); cogs.append(c); gross.append(g)
        op_profit.append(op); net.append(n)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Model"
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)

    for c, h in enumerate(["Metric"] + years, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    rows = [("Revenue", revs), ("COGS", cogs), ("Gross Profit", gross),
            ("Operating Profit", op_profit), ("Net Profit", net)]
    for r, (lbl, vals) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=lbl)
        for c, v in enumerate(vals, 2):
            ws.cell(row=r, column=c, value=round(v, 2))

    ws2 = wb.create_sheet("Assumptions")
    ws2.append(["Parameter", "Value"])
    for k, v in [("Base Revenue", rev0), ("Cost Ratio", cost_ratio),
                  ("Price Increase/yr", price_inc), ("Growth Rate/yr", growth),
                  ("Tax Rate", tax)]:
        ws2.append([k, v])

    wb.save("financial_model.xlsx")

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(years, revs, color="steelblue", label="Revenue", alpha=0.8)
    ax.bar(years, net,  color="green",     label="Net Profit", alpha=0.8)
    ax.set_title("3-Year Financial Projection"); ax.set_ylabel("USD ($)")
    ax.legend(); ax.grid(axis="y", linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.savefig("financial_model.png", dpi=150)
    plt.close()

    summary = "--- 3-Year Financial Model ---\n"
    for i, yr in enumerate(years):
        summary += f"{yr}: Revenue=${revs[i]:,.0f} | Net Profit=${net[i]:,.0f}\n"
    summary += "\nSaved to financial_model.xlsx and financial_model.png."
    return summary


# ── Budget Variance ────────────────────────────────────────────────────────────
@tool
def budget_variance(filename: str) -> str:
    """
    Analyzes budget vs actuals from an uploaded CSV or Excel file.
    The file must contain columns: Category, Budget, Actual.
    """
    path = _find_file(filename)
    if not path:
        return f"Error: '{filename}' not found. Upload the file first."

    try:
        df = _read_file(path)
    except Exception as e:
        return f"Error reading file: {e}"

    df.columns = [c.strip().title() for c in df.columns]
    required = {"Category", "Budget", "Actual"}
    if not required.issubset(set(df.columns)):
        return f"Error: Need columns Category, Budget, Actual. Found: {list(df.columns)}"

    df["Budget"]     = pd.to_numeric(df["Budget"],  errors="coerce")
    df["Actual"]     = pd.to_numeric(df["Actual"],  errors="coerce")
    df["Variance"]   = df["Actual"] - df["Budget"]
    df["Variance %"] = ((df["Variance"] / df["Budget"]) * 100).round(2)
    df["Status"]     = df["Variance"].apply(lambda x: "Over Budget" if x > 0 else "Under Budget")

    df.to_excel("budget_variance.xlsx", index=False)

    fig, ax = plt.subplots(figsize=(12, 5))
    x = range(len(df))
    ax.bar([i - 0.2 for i in x], df["Budget"], width=0.4, label="Budget",  color="steelblue")
    ax.bar([i + 0.2 for i in x], df["Actual"],  width=0.4, label="Actual",  color="coral")
    ax.set_xticks(list(x))
    ax.set_xticklabels(df["Category"], rotation=45, ha="right")
    ax.set_title("Budget vs Actual"); ax.legend()
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.savefig("budget_variance.png", dpi=150)
    plt.close()

    summary = "--- Budget Variance Analysis ---\n"
    for _, row in df.iterrows():
        summary += (f"  {row['Category']}: Budget=${row['Budget']:,.0f} | "
                    f"Actual=${row['Actual']:,.0f} | {row['Variance %']}% ({row['Status']})\n")
    tb, ta = df["Budget"].sum(), df["Actual"].sum()
    summary += (f"\nTotals — Budget: ${tb:,.0f} | Actual: ${ta:,.0f} | "
                f"Net Variance: ${ta-tb:,.0f}\n"
                f"Saved to budget_variance.xlsx and budget_variance.png.")
    return summary


# ── Data Mining ────────────────────────────────────────────────────────────────
@tool
def data_mining(filename: str) -> str:
    """
    Profiles any uploaded CSV or Excel dataset.
    Returns shape, column types, null counts, key stats, and saves a cleaned copy.
    """
    path = _find_file(filename)
    if not path:
        return f"Error: '{filename}' not found. Upload the file first."

    try:
        df = _read_file(path)
    except Exception as e:
        return f"Error reading file: {e}"

    report = f"--- Data Mining Report: {path} ---\n"
    report += f"Shape: {df.shape[0]:,} rows x {df.shape[1]} columns\n\n"
    report += "Column Summary:\n"

    for col in df.columns:
        nulls    = df[col].isnull().sum()
        null_pct = round(nulls / len(df) * 100, 1)
        if pd.api.types.is_numeric_dtype(df[col]):
            report += (f"  {col}: numeric | min={df[col].min():.2f} "
                       f"max={df[col].max():.2f} mean={df[col].mean():.2f} "
                       f"nulls={null_pct}%\n")
        else:
            uniq = df[col].nunique()
            top  = df[col].value_counts().index[0] if not df[col].dropna().empty else "N/A"
            report += f"  {col}: text | {uniq} unique | top='{top}' | nulls={null_pct}%\n"

    dupes    = df.duplicated().sum()
    df_clean = df.dropna(how="all").drop_duplicates()
    clean_path = path.rsplit(".", 1)[0] + "_mined.csv"
    df_clean.to_csv(clean_path, index=False)

    report += (f"\nData Quality: {dupes} duplicate rows | "
               f"{df.isnull().sum().sum()} total nulls\n"
               f"Cleaned file saved to {clean_path}.")
    return report


# ── PDF Report ─────────────────────────────────────────────────────────────────
@tool
def generate_report(ticker: str) -> str:
    """
    Generates a comprehensive PDF report for a stock.
    Requires get_stock_data and clean_and_profile to have been run first.
    """
    from fpdf import FPDF

    ticker     = _clean_ticker(ticker)
    clean_path = f"{ticker}_clean.csv"
    chart_path = f"{ticker}_chart.png"

    if not os.path.exists(clean_path):
        # Try to use raw data if clean doesn't exist
        raw_path = f"{ticker}_raw.csv"
        if not os.path.exists(raw_path):
            return f"Error: No data for {ticker}. Run get_stock_data first."
        df = pd.read_csv(raw_path, index_col=0, parse_dates=True, header=[0, 1])
    else:
        df = pd.read_csv(clean_path, index_col=0, parse_dates=True)

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df = df.apply(pd.to_numeric, errors="coerce")

    # Generate chart if it doesn't exist
    if not os.path.exists(chart_path):
        fig, ax = plt.subplots(figsize=(12, 5))
        ax.plot(df.index, df["Close"], color="royalblue", linewidth=1.5)
        if len(df) >= 50:
            ax.plot(df.index, df["Close"].rolling(50).mean(),
                    color="orange", linewidth=1.2, linestyle="--", label="50-day MA")
        ax.set_title(f"{ticker} — Closing Price")
        ax.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        plt.savefig(chart_path, dpi=150)
        plt.close()

    latest  = round(float(df["Close"].iloc[-1]), 2)
    chg     = round(float(df["Close"].pct_change().iloc[-1]) * 100, 2)
    hi      = round(float(df["Close"].max()), 2)
    lo      = round(float(df["Close"].min()), 2)
    avg     = round(float(df["Close"].mean()), 2)
    std     = round(float(df["Close"].std()), 2)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 12, f"Financial Analysis Report: {ticker}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%B %d, %Y')}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, "Key Statistics", ln=True)
    pdf.set_font("Helvetica", "", 11)

    for lbl, val in [
        ("Latest Close",   f"${latest}"),
        ("1-Day Change",   f"{chg}%"),
        ("52-Week High",   f"${hi}"),
        ("52-Week Low",    f"${lo}"),
        ("Average Close",  f"${avg}"),
        ("Std Deviation",  f"${std}"),
    ]:
        pdf.cell(80, 8, lbl + ":"); pdf.cell(0, 8, val, ln=True)

    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, "Price Chart (Last 1 Year)", ln=True)
    pdf.image(chart_path, x=10, w=190)

    out = f"{ticker}_report.pdf"
    pdf.output(out)
    return f"Success: PDF report saved to {out}."